"""
Author: Georg Börries
Datum: 12.05.2024

Dieses Modul beinhaltet Methoden zum Durchführen der Tests
"""

import os
from datetime import datetime
from multiprocessing import Pool
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.reader.excel import load_workbook

import utilities
from ant_colony_system import AntColonySystem


def pool_optimization_tests(num_coordinates: int, max_distance: int, capacity_drone: int) -> None:
    """
    Diese Funktion führt Tests für die vorgegebene Anzahl an Koordinaten, Distanz und Kapazität der Drone durch

    :param num_coordinates: Anzahl der Koordinaten
    :param max_distance: maximale Entfernung der Koordinaten vom Ursprung
    :param capacity_drone: Kapazität der Drone
    """

    # Koordinaten erzeugen
    depot = (0, 0)  # Koordinaten des Depots auf den Ursprung setzen
    # Koordinaten mit Depot im Zentrum
    delivery_coordinates_center = np.array(
        utilities.generate_delivery_coordinates(depot, num_coordinates, max_distance))
    # Koordinaten mit Depot am Rand
    delivery_coordinates_corner = np.copy(delivery_coordinates_center)  # Kopie erzeugen
    delivery_coordinates_corner[0] = (-max_distance, -max_distance)  # Koordinaten des Depots in eine Ecke setzen

    # Anzahl der Wiederholungen für jeden Test
    n_tests = 3

    # Kapazität des Lieferwagens berechnen in Abhängigkeit von der Kapazität der Drohne und Anzahl der Flüge
    capacities_truck = [num_coordinates - i * capacity_drone for i in range(1, 4)]

    # Liste mit den Parameter-Tupeln für die einzelnen Tests erzeugen
    params = []
    for n in range(1, n_tests + 1):
        for depot in ['center', 'corner']:
            for capacity in capacities_truck:
                for use_drone in [True, False]:
                    params.append((n,
                                   depot,
                                   delivery_coordinates_center if depot == 'center' else delivery_coordinates_corner,
                                   capacity_drone,
                                   capacity,
                                   use_drone))

    # Pool für ACS erzeugen und starten
    pool = Pool(processes = 6)  # Anzahl der gleichzeitigen Prozesse aus Hardwaregründen auf 6 begrenzt
    results = pool.map(pool_acs, params)  # ACS im Pool mit den erzeugten Parameter-Tupeln ausführen
    pool.close()

    # Ordner zum Speichern der Tests
    test_folder = './tests/'
    # Prüfen ob Ordner im Dateisystem vorhanden ist und erstellen falls nicht
    if not os.path.exists(test_folder):
        os.makedirs(test_folder)

    # Zeitpunkt
    date_str = datetime.now().strftime("%Y_%m_%d-%H_%M")
    # Ergebnisse in einem Dataframe speichern
    df_results = pd.DataFrame(results, columns = ['n', 'depot', 'capacity_drone', 'capacity_truck', 'use_drone', 'tour',
                                                  'distance', 'time', 'cost', 'best_iteration', 'fig'])
    # Spalten für die Anzahl der Lieferziele und maximale Distanz zum Dataframe hinzufügen
    df_results['delivery_targets'] = num_coordinates
    df_results['max_distance'] = max_distance

    # Plots für die Touren speichern
    str_hyperlinks = []
    for i, fig in enumerate(df_results['fig']):
        # Dateiname aus den Parametern ableiten
        n = df_results['n'].iloc[i]  # Testlauf
        depot = df_results.loc[i, 'depot']  # Position des Depots
        capacity = df_results.loc[i, 'capacity_truck']  # Kapazität des Lieferwagens
        str_drone = 'with_drone' if df_results.loc[
            i, 'use_drone'] else 'without_drone'  # mit oder ohne Drohnenunterstützung
        # Ordner für die Plots
        figures_folder = (f'{test_folder}figures/drone_capacity_{capacity_drone}/{max_distance}km/'
                          f'{num_coordinates}coordinates/depot_{depot}/{date_str}/')
        # Prüfen ob der Ordner im Dateisystem existiert, sonst erzeugen
        if not os.path.exists(figures_folder):
            os.makedirs(figures_folder)
        # Dateiname
        filename_fig = f'-tour_{str_drone}-capacity_truck{capacity}-{n}.png'
        # Plot speichern
        fig.savefig(figures_folder + filename_fig)
        # Absoluten Pfad auslesen und als Excel-Hyperlink formatieren
        abs_path_fig = Path(os.path.abspath(figures_folder + filename_fig)).as_uri()
        str_hyperlinks.append(f'=HYPERLINK("{abs_path_fig}","Link")')
    # Plots im Dataframe durch die Links zu den Bildern ersetzen
    df_results['fig'] = str_hyperlinks
    # unbearbeitete Ergebnisse als Excel-Datei speichern
    df_results.to_excel(f'./tests/{date_str}-results_raw.xlsx')
    #  benötigte Spalten für die Auswertung im Dataframe filtern
    df_results = df_results[
        ['max_distance', 'delivery_targets', 'depot', 'capacity_drone', 'capacity_truck', 'use_drone',
         'tour', 'distance', 'time', 'cost', 'fig']]
    # neues Dataframe mit den Ergebnissen der Optimierung mit Drohnenunterstützung erzeugen
    df_results_with_drone = pd.DataFrame(df_results[df_results['use_drone'] == True],
                                         columns = ['depot', 'capacity_drone', 'capacity_truck', 'use_drone',
                                                    'tour', 'distance', 'time', 'cost', 'fig'], copy = True)
    # Distanz-Tupel für Lieferwagen, Drohne und kombinierte Distanz in einzelne Spalten aufteilen
    df_results_with_drone[['distance_truck', 'distance_drone', 'distance_combined']] = df_results_with_drone[
        'distance'].tolist()
    # Zeit-Tupel für Lieferwagen und Drohne in einzelne Spalten aufteilen
    df_results_with_drone[['time_truck', 'time_drone']] = df_results_with_drone['time'].tolist()
    # nicht mehr benötigte Spalten entfernen
    df_results_with_drone.drop(columns = ['distance', 'time', 'use_drone', 'capacity_drone'],
                               inplace = True)
    # Spalten umbenennen um doppelte Bezeichnung beim späteren zusammenführen zu Verhinden
    df_results_with_drone.rename(columns = {'tour': '(tour, tour_drone)',
                                            'cost': 'cost_with_drone',
                                            'fig': 'fig_with_drone'},
                                 inplace = True)
    # Aggregation der minimalen Kosten gruppiert nach Kapazität des Lieferwagens und Position des Depots
    min_cost_with_drone = df_results_with_drone.groupby(['capacity_truck', 'depot'])['cost_with_drone'].min()
    # Ergebnisse filtern, sodass nur die Ergebnisse mit den jeweils besten Kosten enthalten sind
    df_results_with_drone = df_results_with_drone[df_results_with_drone['cost_with_drone'].isin(min_cost_with_drone)]
    # neues Dataframe erstellen, dass nur die Ergebnisse der Optimierung ohne Drohnenunterstützung enthält
    df_results_without_drone = df_results[df_results['use_drone'] == False]
    # nicht benötigte Spalten entfernen
    df_results_without_drone = df_results_without_drone.drop(columns = ['use_drone'])
    # Aggregation der minimalen Kosten gruppiert nach Kapazität des Lieferwagens und Position des Depots
    min_cost_without_drone = df_results_without_drone.groupby(['capacity_truck', 'depot'])['cost'].min()
    # Ergebnisse filtern, sodass nur die Ergebnisse mit den minimalen Kosten beibehalten werden
    df_results_without_drone = df_results_without_drone[df_results_without_drone['cost'].isin(min_cost_without_drone)]
    # Indizes zurücksetzen
    df_results_with_drone.reset_index(inplace = True)
    df_results_without_drone.reset_index(inplace = True)
    # neues Dataframe erzeugen mit den kombinierten Ergebnissen
    df_combined_results = df_results_without_drone.merge(df_results_with_drone, on = ['depot', 'capacity_truck'])
    # ggf. Duplikate entfernen
    df_combined_results.drop_duplicates(['cost', 'cost_with_drone'], inplace = True)
    # automatisch erzeugte Index-Spalten entfernen
    df_combined_results.drop(columns = ['index_x', 'index_y'], inplace = True)

    ## Auswertung
    # Differenz der zurückgelegten Distanz des Lieferwagens mit und ohne Drohnenunterstützung berechnen
    df_combined_results['delta_distance'] = df_combined_results['distance'] - df_combined_results['distance_truck']
    # Verhältnis der Differenz zur zurückgelegten Distanz der Drohne berechnen
    df_combined_results['ratio_distance'] = df_combined_results['delta_distance'] / df_combined_results[
        'distance_drone']
    # Differenze der benötigten Zeit des Lieferwagens mit und ohne Drohnenunterstützung berechnen
    df_combined_results['delta_time'] = df_combined_results['time'] - df_combined_results['time_truck']
    # Verhältnis der Differenz zur benötigten Zeit der Drohne berechnen
    df_combined_results['ratio_time'] = df_combined_results['delta_time'] / df_combined_results['time_drone']

    # Spalte mit Durchführungsdatum des Tests einfügen
    df_combined_results['test_date'] = datetime.now()

    # Koordinaten speichern
    folder_coordinates = f'{test_folder}/coordinates/{date_str}/'  # Ordner für die Koordinaten
    if not os.path.exists(folder_coordinates): # prüfen ob der Ordner im Dateisystem existiert, sonst erzeugen
        os.makedirs(folder_coordinates)
    for depot in ['center', 'corner']:
        # Dateiname mit Testdatum und -Zeit sowie Position des Depots
        filename_coordinates = f'{date_str}-coordinates_depot_{depot}.txt'
        # Koordinaten als txt-Datei speichern
        np.savetxt(folder_coordinates + filename_coordinates,
                   delivery_coordinates_center if depot == 'center' else delivery_coordinates_corner)
        # Absoluten Pfad im Dateisystem ermitteln und als Excel-Hyperlink formatiert im Dataframe speichern
        abs_path_coordinates = Path(os.path.abspath(folder_coordinates + filename_coordinates)).as_uri()
        df_combined_results['depot'] = df_combined_results['depot'].replace(depot,
                                                                            f'=HYPERLINK("{abs_path_coordinates}",'
                                                                            f'"{depot}")')

    # Dataframe mit den kombinierten Ergebnissen speichern
    filename = 'combined_results.xlsx'  # Dateiname für die kombinierten Ergebnisse
    if not os.path.exists(test_folder + filename):  # prüfen ob Datei existiert, sonst Dataframe speichern
        df_combined_results.to_excel(test_folder + filename,
                                     sheet_name = 'results',
                                     float_format = '%.2f',
                                     index = False)
    else:   # falls Datei existiert
        with pd.ExcelWriter(test_folder + filename,
                            mode = 'a',
                            if_sheet_exists = 'overlay') as writer:
            # letzte Zeile mit Inhalten ermitteln
            wb = load_workbook(test_folder + filename)
            last_row = wb['results'].max_row
            # Ergebnisse nach der letzten Zeile hinzufügen und speichern
            df_combined_results.to_excel(writer,
                                         sheet_name = 'results',
                                         header = False,
                                         index = False,
                                         startrow = last_row,
                                         float_format = '%.2f')


def pool_acs(params: tuple[int, str, np.ndarray, int, int, bool]) -> tuple[int, str, int, int, bool, any]:
    """
    Funktion zum Aufrufen des ACS als Pool-Funktion, die dynamischen Parameter werden als Tupel übergeben und entpackt
    :param params: Tupel mit den Parametern für Testlauf, Position des Depots, Koordinaten, Kapazität der Drohne,
    Kapazität des Lieferwagens und Optimierung mit/ohne Drohnenunterstützung
    :return:
    """
    # Parameter-Tupel entpacken
    n, depot, coordinates, capacity_drone, capacity_truck, use_drone = params

    # statische ACS Parameter
    max_iterations = 2000
    max_duration = 1800
    alpha = 1
    beta = 2
    ants = 50
    rho = 0.075
    q0 = 0.85

    # statische Fahrzeugparameter
    velocity_truck = 20
    velocity_drone = 50
    cost_per_km_truck = 0.7
    cost_per_km_drone = 0.35
    cost_per_h_truck = 15
    cost_per_h_drone = 20
    loading_time = 0.5 / 60
    delivery_time = 3 / 60

    # ACS initialisieren
    acs = AntColonySystem(coordinates,
                          ants,
                          alpha,
                          beta,
                          rho,
                          q0,
                          max_iterations,
                          max_duration,
                          capacity_truck,
                          capacity_drone,
                          velocity_truck,
                          velocity_drone,
                          cost_per_km_truck,
                          cost_per_km_drone,
                          cost_per_h_truck,
                          cost_per_h_drone,
                          loading_time,
                          delivery_time)

    # Optimierung mit/ohne Drohnenunterstützung durchführen
    if use_drone:
        results = acs.run_truck_with_drone_support(False)
    else:
        results = acs.run_truck_without_drone_support(False)

    return n, depot, capacity_drone, capacity_truck, use_drone, *results


def acs_parameter_tests():
    """
    Funktion zum Ausführen der ACS-Parameter-Tests für q0 und rho
    """
    # Laden des Koordinatensatzes für die Tests
    delivery_coordinates = np.load('tests/Parameter-Tests/coordinates_parameter_test.npy')

    # Fahrzeugparameter
    capacity_truck = 95
    capacity_drone = 3
    velocity_truck = 20
    velocity_drone = 50
    cost_per_km_truck = 0.75
    cost_per_km_drone = 0.3
    cost_per_h_truck = 15
    cost_per_h_drone = 20
    loading_time = 3 / 60
    delivery_time = 3 / 60

    # ACS Parameter
    max_iterations = 3000
    max_duration = 600
    alpha = 1
    beta = 2
    ants = 100
    rho = [0.05, 0.1, 0.15]
    q0 = [0.7, 0.8, 0.9]

    # Listen für die Ergebnisse initialisieren
    results_without_drone = []
    results_with_drone = []

    num_tests = 1
    for _ in range(num_tests):
        for r in rho:
            for q in q0:
                # ACS initialisieren
                acs = AntColonySystem(delivery_coordinates, ants, alpha, beta, r, q, max_iterations, max_duration,
                                      capacity_truck, capacity_drone, velocity_truck, velocity_drone, cost_per_km_truck,
                                      cost_per_km_drone, cost_per_h_truck, cost_per_h_drone, loading_time, delivery_time)
                # Tour ohne Drohnen-Support optimieren
                tour, distance, time, cost, best_iteration = acs.run_truck_without_drone_support()
                results_without_drone.append((ants, alpha, beta, r, q, tour, distance, time, cost, best_iteration))
                # Pheromone zurücksetzen
                acs.reset_pheromones()
                # Tour mit Drohnen-Support optimieren
                tour_drone, distance_drone, time_drone, cost_drone, best_iteration = acs.run_truck_with_drone_support()
                results_with_drone.append(
                    (ants, alpha, beta, r, q, tour_drone, distance_drone, time_drone, cost_drone, best_iteration))

    # Dataframes mit den Ergebnissen mit und ohne Drohnenunterstützung erzeugen
    df_columns = ['ants', 'alpha', 'beta', 'rho', 'q0', 'tour', 'distance', 'time', 'cost', 'best_iteration']
    df_results_without_drone = pd.DataFrame(results_without_drone, columns = df_columns)
    df_results_with_drone = pd.DataFrame(results_with_drone, columns = df_columns)

    # Ergebnisse speichern
    # Dateipfad für Ergebnisse ohne Drohnenunterstützung
    path = './tests/Parameter-Tests/Parameter-Test_ohne_Drohne.xlsx'
    # Prüfen ob Datei existiert und Ergebnisse speichern, falls nicht
    if not os.path.exists(path):
        df_results_without_drone.to_excel(path,
                                          index = False,
                                          sheet_name = 'results')
    else:
        with pd.ExcelWriter(path,
                            mode = 'a',
                            if_sheet_exists = 'overlay') as writer:
            # letzte Zeile mit Inhalten ermitteln
            wb = load_workbook('./results/Parameter-Tests/Parameter-Test_ohne_Drohne.xlsx')
            last_row = wb['results'].max_row
            # Ergebnisse nach der letzten Zeile einfügen und speichern
            df_results_without_drone.to_excel(writer,
                                              sheet_name = 'results',
                                              header = False,
                                              index = False,
                                              startrow = last_row)
    # Dateipfad für Ergebnisse mit Drohnenunterstützung
    path = './tests/Parameter-Tests/Parameter-Test_mit_Drohne.xlsx'
    # Prüfen ob Datei bereits existiert und Ergebnisse speichern, falls nicht
    if not os.path.exists(path):
        df_results_with_drone.to_excel(path,
                                       index = False,
                                       sheet_name = 'results')
    else:
        with pd.ExcelWriter(path,
                            mode = 'a',
                            if_sheet_exists = 'overlay') as writer:
            # letzte Zeile mit inhalten ermitteln
            wb = load_workbook('./results/Parameter-Tests/Parameter-Test_mit_Drohne.xlsx')
            last_row = wb['results'].max_row
            # Ergebnisse nach der letzten Zeile einfügen und speichern
            df_results_with_drone.to_excel(writer,
                                           sheet_name = 'results',
                                           header = False,
                                           index = False,
                                           startrow = last_row)


if __name__ == '__main__':
    for num_coordinates in [50, 100]:
        for max_distance in [5, 10]:
            for capacity_drone in [1, 3, 5]:
                # Optimierungstests für alle möglichen Kombinationen durchführen
                pool_optimization_tests(num_coordinates, max_distance, capacity_drone)
